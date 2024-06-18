import pandas as pd
import numpy as np

from sklearn.model_selection import cross_validate, train_test_split
from sklearn.linear_model import Lasso, Ridge
from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.linear_model import BayesianRidge
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.linear_model import PassiveAggressiveRegressor
from sklearn.gaussian_process.kernels import RBF, ConstantKernel as C
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Dropout
from sklearn.gaussian_process.kernels import RBF, ConstantKernel as C, WhiteKernel  


from I_dataset_creation import visualization_sorted_tab
from II_preprocessing import check_tab, feature_selection_correlation, scale_data, time_series_analysis
from III_train_models import NARX_model, train_arima_model,SANN_model,knn_regressor, linear_regression_sklearn, PolynomialRegression, Neural_Network_Pytorch, ANN_model, random_forest, gradient_boosting_regressor

from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

from statsmodels.tsa.seasonal import seasonal_decompose

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def mean_absolute_percentage_error(y_true, y_pred):
    return np.mean(np.abs((y_true - y_pred) / y_true)) * 100

def Performance(y_test, predictions, name, metrics_data, predictions_data, dates):
    r2 = r2_score(y_test, predictions)
    mse = mean_squared_error(y_test, predictions)
    mae = mean_absolute_error(y_test, predictions)
    mape = mean_absolute_percentage_error(y_test, predictions)

    metrics_data.append({'Model': name,
                         'R^2': round(r2, 4),
                         'MSE': round(mse, 4),
                         'MAE': round(mae, 4),
                         'MAPE': round(mape, 4)})

    # Ensure y_test is a Series or DataFrame
    if isinstance(y_test, pd.Series):
        y_test = y_test.reset_index(drop=True)
    predictions_df = pd.DataFrame({
        'Date': dates,
        'Actual': y_test
    }).reset_index(drop=True)
    
    predictions_df[name] = predictions

    if predictions_data.empty:
        predictions_data = predictions_df
    else:
        predictions_data = pd.merge(predictions_data, predictions_df, on=['Date', 'Actual'], how='left')

    return predictions_data


def dataframe_to_rows(df, index=True, header=True):
    """Converts a pandas DataFrame to rows for writing to an Excel sheet.

    Args:
        df (pd.DataFrame): DataFrame to convert.
        index (bool, optional): Whether to include row labels. Defaults to True.
        header (bool, optional): Whether to include column names. Defaults to True.
    """

    if index:
        yield df.index.tolist()  # Yield row labels if index is True

    if header:
        yield df.columns.tolist()  # Yield column names if header is True

    for row in df.itertuples(index=False, name=None):
        yield list(row)  # Yield each row as a list

def color_predictions_excel(predictions_data):
    """Colors cells in an Excel sheet based on prediction accuracy compared to actual values.

    Args:
        predictions_data (pd.DataFrame): DataFrame containing predictions and actual values.
    """

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the DataFrame to the Excel sheet using the custom dataframe_to_rows function
    for row in dataframe_to_rows(predictions_data, index=False, header=True):
        sheet.append(row)

    # Iterate through columns and rows to apply coloring (rest of the code is the same)
    for col_idx in range(2, predictions_data.shape[1]):
        col_name = predictions_data.columns[col_idx]
        for row_idx in range(1, predictions_data.shape[0] + 1):
            actual_value = predictions_data.iloc[row_idx - 1, 1]
            predicted_value = predictions_data.iloc[row_idx - 1, col_idx]

            if actual_value != 0:
                percentage_diff = abs((predicted_value - actual_value) / actual_value) * 100
            else:
                percentage_diff = 0

            cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)

            if percentage_diff <= 5:
                cell.fill = PatternFill(start_color="03A64A", end_color="03A64A", fill_type="solid")  # Green
            elif percentage_diff <= 10:  
                cell.fill = PatternFill(start_color="F27438", end_color="F27438", fill_type="solid")  # Orange
            else:
                cell.fill = PatternFill(start_color="F24405", end_color="F24405", fill_type="solid")  # Red

    workbook.save("spreadsheet/predictions_colored.xlsx")


def main():
    print("---------------- DATASET CREATION ------------------")

    visualization_sorted_tab()

    data = pd.read_excel('spreadsheet/Data.xlsx')
    colonne_cible = 'Ireland_Milk_Price'

    print(" --------------- PREPROCESSING ------------------")

    # Le tableau est-il exploitable
    exploitable = check_tab(data, colonne_cible)
    print("Le fichier est exploitable :", exploitable)

    past_time = 3
    data = data.iloc[:, 1:]
    data = time_series_analysis(past_time, data)


    feature_selection_correlation

    
    # Corrélation
    # data = correlation_matrix(data, colonne_cible, seuil_corr)

    # Normalisation / standardisation
    # data = scale_data(data, method='normalisation')
    # data = scale_data(data, method='standardisation')


    print("Tableau final :")
    print(data.head())


    print("----------------- TRAIN MODELS -----------------")

    # Split into training and testing sets
    
    data.index = pd.to_datetime(data.index, dayfirst=True)  
    data = data.asfreq(pd.infer_freq(data.index)) 
    X = data.drop(columns=['Ireland_Milk_Price'])
    y = data['Ireland_Milk_Price']
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)


    # Convert y_test to pandas Series with appropriate index
    y_test = pd.Series(y_test.values, index=X_test.index)

    # Collect dates for the test set
    test_dates = X_test.index

    # Vérification des tailles des ensembles
    print(f"Training set size: {len(X_train)}, Test set size: {len(X_test)}")
    print(X_test.head(), X_train.head())

    # List to store the performance metrics
    metrics_data = []
    predictions_data = pd.DataFrame()
    
    # Linear Regression w/ sklearn
    predictions = linear_regression_sklearn(X_train, X_test, y_train)
    predictions_data = Performance(y_test, predictions, 'Linear Regression', metrics_data, predictions_data, test_dates)
    
    # Polynomial regression
    degree = 2
    predictions = PolynomialRegression(X_train, X_test, y_train, degree)
    predictions_data = Performance(y_test, predictions, 'PolynomialRegression', metrics_data, predictions_data, test_dates)

    
    # ANN model w/ tensorflow.keras
    epochs = 100
    batch_size = 10
    validation_split = 0.5
    predictions = ANN_model(X_train, y_train, X_test, epochs, batch_size, validation_split)
    predictions_data = Performance(y_test, predictions, 'ANN_Model', metrics_data, predictions_data, test_dates)

    # Random Forest model
    n_estimators = 100
    random_state = 42
    predictions = random_forest(X_train, X_test, y_train, n_estimators, random_state)
    predictions_data = Performance(y_test, predictions, 'Random Forest', metrics_data, predictions_data, test_dates)

    # Neural Network with PyTorch
    epochs=100
    learning_rate=0.01
    predictions, y_test = Neural_Network_Pytorch(X_train, X_test, y_train, y_test, epochs, learning_rate)
    predictions_data = Performance(y_test, predictions,'Neural Network PyTorch', metrics_data, predictions_data, test_dates)

    # Gradient Boosting Regressor
    n_estimators = 100
    learning_rate = 0.1
    predictions = gradient_boosting_regressor(X_train, X_test, y_train, n_estimators, learning_rate)
    predictions_data = Performance(y_test, predictions, 'Gradient Boosting', metrics_data, predictions_data, test_dates)

    # Lasso Regression
    alpha = 0.1
    lasso = Lasso(alpha=alpha, max_iter=10000) 
    lasso.fit(X_train, y_train)
    predictions = lasso.predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Lasso Regression', metrics_data, predictions_data, test_dates)

    # Ridge Regression
    alpha = 0.5
    ridge = Ridge(alpha=alpha)
    ridge.fit(X_train, y_train)
    predictions = ridge.predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Ridge Regression', metrics_data, predictions_data, test_dates)

    # Support Vector Regressor (SVR)
    svr = SVR(kernel='rbf')
    svr.fit(X_train, y_train)
    predictions = svr.predict(X_test)
    predictions_data = Performance(y_test, predictions, 'SVR', metrics_data, predictions_data, test_dates)

    # KNeighborsRegressor
    n_neighbors = 5
    predictions = knn_regressor(X_train, X_test, y_train, n_neighbors)
    predictions_data = Performance(y_test, predictions, 'KNeighborsRegressor', metrics_data, predictions_data, test_dates)

    # Decision Tree Regressor

    max_depth = 5
    model = DecisionTreeRegressor(max_depth=max_depth)
    predictions = model.fit(X_train, y_train).predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Decision Tree Regressor', metrics_data, predictions_data, test_dates)

    # Bayesian Ridge
    
    model = BayesianRidge()
    predictions = model.fit(X_train, y_train).predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Bayesian Ridge', metrics_data, predictions_data, test_dates)

    # Extra Trees Regressor

    n_estimators = 100
    model = ExtraTreesRegressor(n_estimators=n_estimators)
    predictions = model.fit(X_train, y_train).predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Extra Trees Regressor', metrics_data, predictions_data, test_dates)

    # Passive Aggressive Regressor

    model = PassiveAggressiveRegressor()
    predictions = model.fit(X_train, y_train).predict(X_test)
    predictions_data = Performance(y_test, predictions, 'Passive Aggressive Regressor', metrics_data, predictions_data, test_dates)


    # NARX (Nonlinear AutoRegressive with eXogenous inputs)
    predictions = NARX_model(X_train, y_train, X_test)
    predictions_data = Performance(y_test, predictions, 'NARX', metrics_data, predictions_data, test_dates)


    """
    # SANN (Self-Adaptive Neural Network)
    # Utilisation du modèle SANN
    epochs = 50
    batch_size = 16
    predictions = SANN_model(X_train, y_train, X_test, epochs, batch_size)
    predictions_data = Performance(y_test, predictions, 'SANN', metrics_data, predictions_data, test_dates)

    
    
    # ARIMA model
    data.index = pd.to_datetime(data.index, dayfirst=True)  
    data = data.asfreq(pd.infer_freq(data.index)) 
    X = data.drop(columns=['Ireland_Milk_Price'])
    y = data['Ireland_Milk_Price']
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, shuffle=False)
    y_test = pd.Series(y_test.values, index=X_test.index)

    predictions = train_arima_model(y_train, y_test)
    predictions_data = Performance(y_test, predictions, 'ARIMA', metrics_data, predictions_data, test_dates)"""


    # Export metrics
    metrics_df = pd.DataFrame(metrics_data)
    metrics_df = metrics_df.sort_values(by='MAPE', ascending=True)
    metrics_df.to_excel('spreadsheet/emetrics.xlsx', index=False)

    # Export combined predictions
    predictions_data.to_excel('spreadsheet/predictions.xlsx', index=False)
    df_predictions_rounded = predictions_data.copy()  # Create a copy to avoid modifying the original
    df_predictions_rounded.iloc[:, 1:] = df_predictions_rounded.iloc[:, 1:].round(3)


    # Call the function with the DataFrame
    color_predictions_excel(df_predictions_rounded)




if __name__ == '__main__':
    main()
